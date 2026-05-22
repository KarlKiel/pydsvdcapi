#!/usr/bin/env python3
"""Generate full_showcase_devices.xlsx from live device instances.

Instantiates all 27 showcase devices, calls get_properties() on each, and
writes the actual announced values to Excel.  Every API-defined property key
is listed even when not set; a column marks each as Required or Optional.
Sheet 1 covers the VDC itself; sheets 2–28 cover each vdSD.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from full_showcase import STATE_FILE, VENDOR_NAME, build_all_devices
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pydsvdcapi import Vdc, VdcCapabilities, VdcHost

OUT_PATH = Path(__file__).parent / "full_showcase_devices.xlsx"

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
C_DEV_TITLE = "1F3864"  # device title bar
C_SECTION = "2E75B6"  # section header
C_COL_HEADER = "4472C4"  # column label row
C_KEY = "D6E4F7"  # property key cell
C_SUBKEY = "EBF3FB"  # sub-key cell (indent level 2)
C_VAL_A = "FFFFFF"  # value, odd row
C_VAL_B = "F5F9FF"  # value, even row
C_REQ = "E2EFDA"  # required  → light green
C_OPT = "FFF2CC"  # optional  → light yellow
C_OPT_REC = "FCE4D6"  # optional* (recommended) → light orange

_GROUP_TITLE_FILL = {
    1: "FFD700",
    2: "C0C0C0",
    3: "87CEEB",
    4: "E0FFFF",
    5: "EE82EE",
    6: "FF6B6B",
    7: "90EE90",
    8: "A9A9A9",
    9: "F0F0F0",
    10: "B0E0E6",
    11: "ADD8E6",
    12: "87CEEB",
    48: "AFEEEE",
    64: "E0FFFF",
    65: "D3D3D3",
}

_SENTINEL = object()  # marks "key not present in actual props"

# ---------------------------------------------------------------------------
# Complete API property schemas
# Each entry: (key, required_label, note)
# required_label: "Required" | "Optional" | "Optional*"
# ---------------------------------------------------------------------------

SCHEMA_IDENT_VDSD = [
    ("dSUID", "Required", "Auto-generated unique device identifier"),
    ("displayId", "Optional", "Human-readable short display ID"),
    ("type", "Required", "Entity type string (auto-set by library)"),
    ("name", "Required", "Human-readable device name"),
    ("model", "Required", "Device model identifier"),
    ("modelVersion", "Optional", "Semantic version string e.g. '1.0.0'"),
    ("modelUID", "Optional", "Unique model type ID (auto-derived from GTIN if set)"),
    ("hardwareVersion", "Optional", "Hardware revision string"),
    (
        "hardwareGuid",
        "Optional*",
        "Unique instance ID mac-address:XX:…; required for GTIN and name distribution",
    ),
    (
        "hardwareModelGuid",
        "Optional*",
        "Model-class ID ean:(01)…; required for GTIN and name distribution",
    ),
    ("vendorName", "Optional", "Manufacturer display name"),
    ("vendorId", "Optional", "Numeric vendor identifier"),
    ("vendorGuid", "Optional", "Global vendor GUID gs1:(01)…"),
    ("descriptionsGroup", "Optional", "Descriptions group code"),
    ("descriptionsClass", "Optional", "Descriptions class code"),
    ("oemGuid", "Optional", "OEM manufacturer GUID"),
    (
        "oemModelGuid",
        "Optional",
        "GTIN / OEM model GUID gs1:(01)…; enables hasActions in dSS VdcDb",
    ),
    ("configURL", "Optional", "URL for custom device configuration page"),
    ("deviceIcon16", "Optional", "16×16 device icon (base64-encoded PNG)"),
    ("deviceIconName", "Optional", "Icon name from built-in dSS icon set"),
    ("deviceClass", "Optional", "Device class code"),
    ("deviceClassVersion", "Optional", "Device class version"),
    ("active", "Required", "Whether device is currently active (auto-set)"),
    (
        "primaryGroup",
        "Required",
        "Device colour / application type (ColorClass enum integer)",
    ),
    ("zoneID", "Optional", "Zone ID; 0 = unassigned to a zone"),
    ("progMode", "Optional", "Programming mode flag"),
    ("currentConfigId", "Optional", "Active configuration profile ID"),
]

SCHEMA_IDENT_VDC = [
    ("dSUID", "Required", "Auto-generated unique VDC identifier"),
    ("displayId", "Optional", "Human-readable short display ID"),
    ("type", "Required", "Entity type string (auto-set by library)"),
    ("name", "Required", "VDC display name"),
    ("model", "Required", "VDC model string"),
    ("modelVersion", "Optional", "Semantic version string"),
    ("modelUID", "Optional", "Unique model type identifier"),
    ("hardwareVersion", "Optional", "Hardware revision string"),
    ("hardwareGuid", "Optional", "Unique hardware instance ID"),
    ("hardwareModelGuid", "Optional", "Hardware model class ID"),
    ("vendorName", "Optional", "Manufacturer display name"),
    ("vendorId", "Optional", "Numeric vendor ID"),
    ("vendorGuid", "Optional", "Global vendor GUID"),
    ("descriptionsGroup", "Optional", "Descriptions group"),
    ("descriptionsClass", "Optional", "Descriptions class"),
    ("oemGuid", "Optional", "OEM manufacturer GUID"),
    ("oemModelGuid", "Optional", "OEM model GUID / GTIN"),
    ("configURL", "Optional", "Configuration page URL"),
    ("deviceIcon16", "Optional", "16×16 icon (base64 PNG)"),
    ("deviceIconName", "Optional", "Built-in icon name"),
    ("deviceClass", "Optional", "Device class code"),
    ("deviceClassVersion", "Optional", "Device class version"),
    ("active", "Required", "Active flag (auto-set)"),
    ("implementationId", "Required", "Unique VDC implementation identifier string"),
    ("zoneID", "Optional", "Zone ID (default: 0)"),
]

SCHEMA_VDC_CAPABILITIES = [
    ("metering", "Optional", "Energy metering support (default: false)"),
    (
        "identification",
        "Optional",
        "Device identification / blink support (default: false)",
    ),
    (
        "dynamicDefinitions",
        "Optional",
        "dSS queries VDC live for states/events/actions (default: false)",
    ),
]

SCHEMA_MODEL_FEATURES = []  # handled specially — just a flat set of names

SCHEMA_OUTPUT_DESC = [
    (
        "function",
        "Required",
        "OutputFunction: 0=ON_OFF 1=POSITIONAL 2=DIMMER 3=DIMMER_COLOR_TEMP 4=FULL_COLOR_DIMMER 5=INTERNALLY_CONTROLLED 6=BIPOLAR 255=CUSTOM",
    ),
    ("outputUsage", "Required", "OutputUsage: 0=UNDEFINED 1=ROOM 2=OUTDOORS"),
    ("name", "Optional", "Output name string"),
    ("defaultGroup", "Required", "Default scene group ID"),
    ("variableRamp", "Optional", "True if ramp time is configurable (default: false)"),
    ("maxPower", "Optional", "Maximum rated power in Watts"),
    (
        "activeCoolingMode",
        "Optional",
        "True for dual heat/cool outputs (climate devices)",
    ),
]

SCHEMA_OUTPUT_SETTINGS = [
    ("mode", "Required", "OutputMode: 0=DISABLED 2=GRADUAL 35=BINARY 127=DEFAULT"),
    ("activeGroup", "Required", "Currently active scene group ID"),
    (
        "pushChanges",
        "Required",
        "Push value changes to dSS proactively (true recommended)",
    ),
    ("groups", "Required", "Scene group membership {groupId: true, …}"),
    ("onThreshold", "Optional", "Value above which output is 'on' (%)"),
    ("minBrightness", "Optional", "Minimum non-zero brightness (%)"),
    ("dimTimeUp", "Optional", "Ramp-up time (ms)"),
    ("dimTimeDown", "Optional", "Ramp-down time (ms)"),
    ("dimTimeUpAlt1", "Optional", "Alternative ramp-up 1 (ms)"),
    ("dimTimeDownAlt1", "Optional", "Alternative ramp-down 1 (ms)"),
    ("dimTimeUpAlt2", "Optional", "Alternative ramp-up 2 (ms)"),
    ("dimTimeDownAlt2", "Optional", "Alternative ramp-down 2 (ms)"),
    (
        "heatingSystemCapability",
        "Optional",
        "HeatingSystemCapability enum — required for climate outputs",
    ),
    (
        "heatingSystemType",
        "Optional",
        "HeatingSystemType enum — required for climate outputs",
    ),
]

SCHEMA_OUTPUT_STATE = [
    ("localPriority", "Required", "Local priority override active"),
    ("error", "Required", "Output error code (0 = no error)"),
]

SCHEMA_CHANNEL_DESC = [
    ("name", "Required", "Channel name string e.g. 'brightness'"),
    ("channelType", "Required", "OutputChannelType enum value"),
    ("dsIndex", "Required", "Zero-based channel index"),
    ("min", "Required", "Minimum physical value"),
    ("max", "Required", "Maximum physical value"),
    ("resolution", "Required", "Smallest increment (physical units)"),
]

SCHEMA_CHANNEL_SETTINGS = []  # no per-channel settings defined in current API

SCHEMA_CHANNEL_STATE = [
    ("value", "Required", "Current channel value (null = not yet known)"),
    ("age", "Required", "Seconds since last update (null = never)"),
]

SCHEMA_BTN_DESC = [
    ("name", "Required", "Button name string"),
    ("dsIndex", "Required", "Zero-based element index"),
    ("supportsLocalKeyMode", "Required", "Whether local key mode is supported"),
    (
        "buttonType",
        "Required",
        "ButtonType: 0=UNDEFINED 1=SINGLE_PUSHBUTTON 2=TWO_WAY 3=FOUR_WAY_NAV 4=FOUR_WAY_CENTER 6=ON_OFF_SWITCH",
    ),
    (
        "buttonElementID",
        "Required",
        "ButtonElementID: 0=CENTER 1=DOWN 2=UP 3=LEFT 4=RIGHT …",
    ),
    ("buttonID", "Optional", "Groups elements of the same physical button unit"),
]

SCHEMA_BTN_SETTINGS = [
    ("group", "Required", "Scene group the button acts on"),
    ("function", "Required", "ButtonFunction: 0=DEVICE 1-4=AREA 5=ROOM 15=APP …"),
    ("mode", "Required", "ButtonMode: 0=STANDARD 1=TURBO …"),
    ("channel", "Optional", "Output channel index to control (for DEVICE function)"),
    (
        "setsLocalPriority",
        "Optional",
        "Whether pressing sets local priority (default: false)",
    ),
    (
        "callsPresent",
        "Optional",
        "Whether pressing sends 'present' call (default: false)",
    ),
]

SCHEMA_BTN_STATE = [
    ("value", "Required", "Last click value (click mode; null = never pressed)"),
    ("clickType", "Required", "ClickType enum (click mode)"),
    ("age", "Required", "Seconds since last press (null = never)"),
    ("error", "Required", "Error code (0 = no error)"),
    (
        "actionId",
        "Optional",
        "Action ID — used in action mode instead of value/clickType",
    ),
    ("actionMode", "Optional", "ActionMode enum — used in action mode"),
]

SCHEMA_BINARY_DESC = [
    ("name", "Required", "Input name string"),
    ("dsIndex", "Required", "Zero-based index"),
    ("inputType", "Optional", "Hardware input type code"),
    ("inputUsage", "Required", "BinaryInputUsage enum"),
    (
        "sensorFunction",
        "Required",
        "BinaryInputType enum (hardwired sensor classification)",
    ),
    ("updateInterval", "Optional", "Expected update interval in seconds"),
]

SCHEMA_BINARY_SETTINGS = [
    ("group", "Required", "Scene group for automation rules"),
    ("sensorFunction", "Required", "BinaryInputType enum (user-configurable override)"),
]

SCHEMA_BINARY_STATE = [
    ("value", "Required", "Current boolean state (null = unknown)"),
    ("extendedValue", "Optional", "Extended state value (replaces value when set)"),
    ("age", "Required", "Seconds since last change (null = never)"),
    ("error", "Required", "Error code (0 = no error)"),
]

SCHEMA_SENSOR_DESC = [
    ("name", "Required", "Sensor name string"),
    ("dsIndex", "Required", "Zero-based index"),
    ("sensorType", "Required", "SensorType enum value"),
    ("sensorUsage", "Required", "SensorUsage enum value"),
    ("min", "Required", "Minimum measurable value"),
    ("max", "Required", "Maximum measurable value"),
    ("resolution", "Required", "Measurement resolution"),
    ("updateInterval", "Optional", "Expected update interval in seconds"),
    ("aliveSignInterval", "Optional", "Alive-sign interval in seconds"),
]

SCHEMA_SENSOR_SETTINGS = [
    ("group", "Required", "Scene group for automation rules"),
    ("minPushInterval", "Optional", "Minimum push interval in seconds"),
    (
        "changesOnlyInterval",
        "Optional",
        "Push on change only within this interval (seconds)",
    ),
]

SCHEMA_SENSOR_STATE = [
    ("value", "Required", "Current sensor reading (null = unknown)"),
    ("age", "Required", "Seconds since last measurement (null = never)"),
    ("contextId", "Optional", "Measurement context identifier"),
    ("contextMsg", "Optional", "Context message string"),
]

SCHEMA_ACTION_DESC = [
    ("name", "Required", "Action identifier string"),
    ("params", "Optional", "Parameter descriptors {paramName: {type, default, …}}"),
    ("description", "Optional", "Human-readable description"),
]

SCHEMA_STD_ACTION = [
    ("name", "Required", "Standard action name e.g. 'std.play'"),
    ("action", "Required", "Target action ID"),
    ("params", "Optional", "Fixed parameter values {paramName: value}"),
]

SCHEMA_EVENT_DESC = [
    ("name", "Required", "Event identifier string"),
    ("description", "Optional", "Human-readable description"),
]

SCHEMA_STATE_DESC = [
    ("name", "Required", "State identifier string"),
    ("value", "Required", "Enum options dict {values: {label: NO_VALUE, …}}"),
    ("description", "Optional", "Human-readable description"),
]

SCHEMA_STATE_VAL = [
    ("name", "Required", "State identifier (matches description key)"),
    ("value", "Required", "Current state label string (null = not yet set)"),
]

SCHEMA_PROP_DESC = [
    ("name", "Required", "Property identifier string"),
    ("type", "Required", "'numeric' | 'string' | 'boolean' | 'enumeration'"),
    ("min", "Optional", "Minimum value (numeric only)"),
    ("max", "Optional", "Maximum value (numeric only)"),
    ("resolution", "Optional", "Step increment (numeric only)"),
    ("siunit", "Optional", "SI unit string e.g. '°C', 's' (numeric only)"),
    ("values", "Optional", "Enum label dict (enumeration only)"),
    ("default", "Optional", "Default value"),
    ("description", "Optional", "Human-readable description"),
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(wrap: bool = True) -> Alignment:
    return Alignment(wrap_text=wrap, vertical="top")


def _req_fill(label: str) -> str:
    if label == "Required":
        return C_REQ
    if label == "Optional*":
        return C_OPT_REC
    return C_OPT


def _format_val(v) -> str:
    if v is _SENTINEL:
        return "—  (not set)"
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float):
        return f"{v:.6g}"
    if isinstance(v, dict):
        if not v:
            return "{}"
        return " | ".join(f"{k}: {_format_val(val)}" for k, val in v.items())
    if isinstance(v, (list, set)):
        return ", ".join(str(x) for x in v)
    return str(v)


# ---------------------------------------------------------------------------
# Sheet-level helpers
# ---------------------------------------------------------------------------


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_title_row(
    ws, row: int, text: str, fill_color: str = C_DEV_TITLE, ncols: int = 5
) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = _fill(fill_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border()
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_col_headers(
    ws,
    row: int,
    labels: tuple = ("Property", "Sub-key / Index", "Value", "Req / Opt", "API Note"),
) -> int:
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(C_COL_HEADER)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_section_header(ws, row: int, text: str, ncols: int = 5) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = _fill(C_SECTION)
    c.border = _border()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 15
    return row + 1


def _write_index_banner(ws, row: int, text: str, ncols: int = 5) -> int:
    """Sub-header for indexed items like 'Channel [0]'."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"    {text}")
    c.font = Font(bold=True, italic=True, size=9)
    c.fill = _fill(C_KEY)
    c.border = _border()
    c.alignment = _align(wrap=False)
    return row + 1


def _write_schema_rows(
    ws, row: int, schema: list, actual: dict, indent: int = 0
) -> int:
    """Write one row per schema entry.  Looks up actual value; shows sentinel if absent."""
    alt = False
    for key, req_label, note in schema:
        val = actual.get(key, _SENTINEL)
        val_str = _format_val(val)
        val_bg = C_VAL_B if alt else C_VAL_A
        req_bg = _req_fill(req_label)

        indent_str = "  " * indent

        # Col A – property key
        ka = ws.cell(row=row, column=1, value=f"{indent_str}{key}")
        ka.font = Font(size=9)
        ka.fill = _fill(C_KEY)
        ka.border = _border()
        ka.alignment = _align()

        # Col B – sub-key (empty at this level)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
        kb = ws.cell(row=row, column=2, value="")
        kb.fill = _fill(C_SUBKEY)
        kb.border = _border()

        # Col C – value
        vc = ws.cell(row=row, column=3, value=val_str)
        vc.font = Font(size=9)
        vc.fill = _fill(val_bg)
        vc.border = _border()
        vc.alignment = _align()

        # Col D – required / optional
        dc = ws.cell(row=row, column=4, value=req_label)
        dc.font = Font(size=9, bold=(req_label == "Required"))
        dc.fill = _fill(req_bg)
        dc.border = _border()
        dc.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # Col E – note
        nc = ws.cell(row=row, column=5, value=note)
        nc.font = Font(size=8, italic=True, color="444444")
        nc.fill = _fill(val_bg)
        nc.border = _border()
        nc.alignment = _align()

        row += 1
        alt = not alt
    return row


def _write_nested_schema_rows(
    ws, row: int, outer_key: str, schema: list, actual: dict
) -> int:
    """Write schema rows with outer_key in Col A and sub-key in Col B."""
    alt = False
    first = True
    for key, req_label, note in schema:
        val = actual.get(key, _SENTINEL)
        val_str = _format_val(val)
        val_bg = C_VAL_B if alt else C_VAL_A
        req_bg = _req_fill(req_label)

        ka = ws.cell(row=row, column=1, value=outer_key if first else "")
        ka.font = Font(size=9)
        ka.fill = _fill(C_KEY)
        ka.border = _border()
        ka.alignment = _align()
        first = False

        kb = ws.cell(row=row, column=2, value=key)
        kb.font = Font(size=9, italic=True)
        kb.fill = _fill(C_SUBKEY)
        kb.border = _border()
        kb.alignment = _align()

        vc = ws.cell(row=row, column=3, value=val_str)
        vc.font = Font(size=9)
        vc.fill = _fill(val_bg)
        vc.border = _border()
        vc.alignment = _align()

        dc = ws.cell(row=row, column=4, value=req_label)
        dc.font = Font(size=9, bold=(req_label == "Required"))
        dc.fill = _fill(req_bg)
        dc.border = _border()
        dc.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        nc = ws.cell(row=row, column=5, value=note)
        nc.font = Font(size=8, italic=True, color="444444")
        nc.fill = _fill(val_bg)
        nc.border = _border()
        nc.alignment = _align()

        row += 1
        alt = not alt
    return row


# ---------------------------------------------------------------------------
# VDC sheet
# ---------------------------------------------------------------------------


def build_vdc_sheet(wb: openpyxl.Workbook, vdc_props: dict) -> None:
    ws = wb.create_sheet(title="00 – VDC")
    _set_col_widths(ws, [28, 22, 40, 14, 50])

    row = _write_title_row(ws, 1, "VDC — Virtual Device Controller", C_DEV_TITLE)
    row += 1
    row = _write_col_headers(ws, row)

    # Identification
    row = _write_section_header(ws, row, "IDENTIFICATION")
    row = _write_schema_rows(ws, row, SCHEMA_IDENT_VDC, vdc_props)
    row += 1

    # Capabilities (nested dict)
    caps_actual = vdc_props.get("capabilities", {})
    row = _write_section_header(ws, row, "CAPABILITIES")
    row = _write_nested_schema_rows(
        ws, row, "capabilities", SCHEMA_VDC_CAPABILITIES, caps_actual
    )
    row += 1

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# vdSD sheet
# ---------------------------------------------------------------------------


def build_vdsd_sheet(
    wb: openpyxl.Workbook, di_idx: int, name: str, props: dict, primary_group: int
) -> None:
    safe = (
        name.replace("/", "-")
        .replace("\\", "-")
        .replace("?", "")
        .replace("*", "")
        .replace("[", "(")
        .replace("]", ")")
        .replace(":", "-")
    )
    ws = wb.create_sheet(title=f"D{di_idx:02d} {safe}"[:31])
    _set_col_widths(ws, [28, 22, 40, 14, 50])

    _GROUP_TITLE_FILL.get(primary_group, "CCCCCC")
    row = _write_title_row(
        ws,
        1,
        f"D{di_idx:02d}  {name}  ·  primaryGroup = {primary_group}",
        fill_color=C_DEV_TITLE,
    )
    row += 1
    row = _write_col_headers(ws, row)

    # ------------------------------------------------------------------ #
    # 1. Identification
    # ------------------------------------------------------------------ #
    row = _write_section_header(ws, row, "IDENTIFICATION")
    row = _write_schema_rows(ws, row, SCHEMA_IDENT_VDSD, props)
    row += 1

    # ------------------------------------------------------------------ #
    # 2. Model Features
    # ------------------------------------------------------------------ #
    mf = props.get("modelFeatures", {})
    row = _write_section_header(ws, row, "MODEL FEATURES")
    if mf:
        all_features = sorted(mf.keys())
        # One row per feature
        for i, feat in enumerate(all_features):
            val_bg = C_VAL_B if i % 2 else C_VAL_A
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ka = ws.cell(row=row, column=1, value=feat)
            ka.font = Font(size=9)
            ka.fill = _fill(C_KEY)
            ka.border = _border()
            ka.alignment = _align()
            vc = ws.cell(row=row, column=3, value="true")
            vc.font = Font(size=9)
            vc.fill = _fill(val_bg)
            vc.border = _border()
            dc = ws.cell(row=row, column=4, value="Optional")
            dc.font = Font(size=9)
            dc.fill = _fill(C_OPT)
            dc.border = _border()
            dc.alignment = Alignment(horizontal="center", vertical="top")
            nc = ws.cell(
                row=row,
                column=5,
                value="One entry per enabled feature flag (boolean true)",
            )
            nc.font = Font(size=8, italic=True, color="444444")
            nc.fill = _fill(val_bg)
            nc.border = _border()
            nc.alignment = _align()
            row += 1
    else:
        row = _write_schema_rows(
            ws,
            row,
            [
                (
                    "modelFeatures",
                    "Optional",
                    "Dict of enabled feature flags {featureName: true}",
                )
            ],
            props,
        )
    row += 1

    # ------------------------------------------------------------------ #
    # 3. Output
    # ------------------------------------------------------------------ #
    if "outputDescription" in props or "outputSettings" in props:
        row = _write_section_header(ws, row, "OUTPUT")
        out_desc = props.get("outputDescription", {})
        out_set = props.get("outputSettings", {})
        out_st = props.get("outputState", {})
        row = _write_nested_schema_rows(
            ws, row, "outputDescription", SCHEMA_OUTPUT_DESC, out_desc
        )
        row = _write_nested_schema_rows(
            ws, row, "outputSettings", SCHEMA_OUTPUT_SETTINGS, out_set
        )
        row = _write_nested_schema_rows(
            ws, row, "outputState", SCHEMA_OUTPUT_STATE, out_st
        )
        row += 1

    # ------------------------------------------------------------------ #
    # 4. Channels
    # ------------------------------------------------------------------ #
    ch_desc = props.get("channelDescriptions", {})
    ch_set = props.get("channelSettings", {})
    ch_state = props.get("channelStates", {})
    if ch_desc or ch_set or ch_state:
        row = _write_section_header(ws, row, "CHANNELS")
        # Channel keys are names (e.g. "brightness"); sort by dsIndex for display.
        all_names = sorted(
            set(list(ch_desc) + list(ch_set) + list(ch_state)),
            key=lambda n: ch_desc.get(n, {}).get("dsIndex", 0),
        )
        for ch_name in all_names:
            row = _write_index_banner(ws, row, f"Channel  —  {ch_name}")
            row = _write_nested_schema_rows(
                ws, row, "  description", SCHEMA_CHANNEL_DESC, ch_desc.get(ch_name, {})
            )
            if SCHEMA_CHANNEL_SETTINGS:
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  settings",
                    SCHEMA_CHANNEL_SETTINGS,
                    ch_set.get(ch_name, {}),
                )
            row = _write_nested_schema_rows(
                ws, row, "  state", SCHEMA_CHANNEL_STATE, ch_state.get(ch_name, {})
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 5. Scenes (summary only — hundreds of rows otherwise)
    # ------------------------------------------------------------------ #
    scenes = props.get("scenes", {})
    if scenes:
        row = _write_section_header(ws, row, f"SCENES  ({len(scenes)} entries)")
        nums = sorted(scenes.keys(), key=int)
        preview = ", ".join(nums[:12]) + (
            f"  … ({len(nums)} total)" if len(nums) > 12 else ""
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ka = ws.cell(row=row, column=1, value="scene numbers present")
        ka.font = Font(size=9)
        ka.fill = _fill(C_KEY)
        ka.border = _border()
        ka.alignment = _align()
        vc = ws.cell(row=row, column=3, value=preview)
        vc.font = Font(size=9)
        vc.fill = _fill(C_VAL_A)
        vc.border = _border()
        vc.alignment = _align()
        dc = ws.cell(row=row, column=4, value="Required")
        dc.font = Font(size=9, bold=True)
        dc.fill = _fill(C_REQ)
        dc.border = _border()
        dc.alignment = Alignment(horizontal="center", vertical="top")
        nc = ws.cell(
            row=row,
            column=5,
            value="Scene value table — Required when output has channels",
        )
        nc.font = Font(size=8, italic=True, color="444444")
        nc.fill = _fill(C_VAL_A)
        nc.border = _border()
        nc.alignment = _align()
        row += 1
        row += 1

    # ------------------------------------------------------------------ #
    # 6. Button Inputs
    # ------------------------------------------------------------------ #
    btn_desc = props.get("buttonInputDescriptions", {})
    btn_set = props.get("buttonInputSettings", {})
    btn_st = props.get("buttonInputStates", {})
    if btn_desc or btn_set or btn_st:
        row = _write_section_header(ws, row, "BUTTON INPUTS")
        all_idx = sorted(set(list(btn_desc) + list(btn_set) + list(btn_st)), key=int)
        for idx_str in all_idx:
            bname = btn_desc.get(idx_str, {}).get("name", idx_str)
            row = _write_index_banner(ws, row, f"Button [{idx_str}]  —  {bname}")
            row = _write_nested_schema_rows(
                ws, row, "  description", SCHEMA_BTN_DESC, btn_desc.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  settings", SCHEMA_BTN_SETTINGS, btn_set.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  state", SCHEMA_BTN_STATE, btn_st.get(idx_str, {})
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 7. Binary Inputs
    # ------------------------------------------------------------------ #
    bi_desc = props.get("binaryInputDescriptions", {})
    bi_set = props.get("binaryInputSettings", {})
    bi_st = props.get("binaryInputStates", {})
    if bi_desc or bi_set or bi_st:
        row = _write_section_header(ws, row, "BINARY INPUTS")
        all_idx = sorted(set(list(bi_desc) + list(bi_set) + list(bi_st)), key=int)
        for idx_str in all_idx:
            bname = bi_desc.get(idx_str, {}).get("name", idx_str)
            row = _write_index_banner(ws, row, f"Binary Input [{idx_str}]  —  {bname}")
            row = _write_nested_schema_rows(
                ws, row, "  description", SCHEMA_BINARY_DESC, bi_desc.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  settings", SCHEMA_BINARY_SETTINGS, bi_set.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  state", SCHEMA_BINARY_STATE, bi_st.get(idx_str, {})
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 8. Sensor Inputs
    # ------------------------------------------------------------------ #
    si_desc = props.get("sensorDescriptions", {})
    si_set = props.get("sensorSettings", {})
    si_st = props.get("sensorStates", {})
    if si_desc or si_set or si_st:
        row = _write_section_header(ws, row, "SENSOR INPUTS")
        all_idx = sorted(set(list(si_desc) + list(si_set) + list(si_st)), key=int)
        for idx_str in all_idx:
            sname = si_desc.get(idx_str, {}).get("name", idx_str)
            row = _write_index_banner(ws, row, f"Sensor [{idx_str}]  —  {sname}")
            row = _write_nested_schema_rows(
                ws, row, "  description", SCHEMA_SENSOR_DESC, si_desc.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  settings", SCHEMA_SENSOR_SETTINGS, si_set.get(idx_str, {})
            )
            row = _write_nested_schema_rows(
                ws, row, "  state", SCHEMA_SENSOR_STATE, si_st.get(idx_str, {})
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 9. SingleDevice — Action Descriptions
    # ------------------------------------------------------------------ #
    act_desc = props.get("deviceActionDescriptions")
    if act_desc is not None:
        row = _write_section_header(
            ws, row, "ACTION DESCRIPTIONS  (SingleDevice §4.5.2)"
        )
        if act_desc:
            for name_key, entry in act_desc.items():
                row = _write_index_banner(ws, row, f"Action  '{name_key}'")
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  description",
                    SCHEMA_ACTION_DESC,
                    entry if isinstance(entry, dict) else {},
                )
        else:
            row = _write_schema_rows(
                ws,
                row,
                [
                    (
                        "(empty — no actions)",
                        "Required",
                        "Must be present for SingleDevice even if empty",
                    )
                ],
                {},
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 10. Standard Actions
    # ------------------------------------------------------------------ #
    std_act = props.get("standardActions")
    if std_act is not None:
        row = _write_section_header(ws, row, "STANDARD ACTIONS  (SingleDevice §4.5.3)")
        if std_act:
            for name_key, entry in std_act.items():
                row = _write_index_banner(ws, row, f"Standard Action  '{name_key}'")
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  properties",
                    SCHEMA_STD_ACTION,
                    entry if isinstance(entry, dict) else {},
                )
        else:
            row = _write_schema_rows(ws, row, [("(empty)", "Optional", "")], {})
        row += 1

    # ------------------------------------------------------------------ #
    # 11. Event Descriptions
    # ------------------------------------------------------------------ #
    ev_desc = props.get("deviceEventDescriptions")
    if ev_desc is not None:
        row = _write_section_header(
            ws, row, "EVENT DESCRIPTIONS  (SingleDevice §4.7.1)"
        )
        if ev_desc:
            for name_key, entry in ev_desc.items():
                row = _write_index_banner(ws, row, f"Event  '{name_key}'")
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  description",
                    SCHEMA_EVENT_DESC,
                    entry if isinstance(entry, dict) else {},
                )
        else:
            row = _write_schema_rows(
                ws,
                row,
                [("(empty)", "Required", "Must be present for SingleDevice")],
                {},
            )
        row += 1

    # ------------------------------------------------------------------ #
    # 12. State Descriptions + State Values
    # ------------------------------------------------------------------ #
    st_desc = props.get("deviceStateDescriptions")
    st_vals = props.get("deviceStates", {})
    if st_desc is not None:
        row = _write_section_header(
            ws, row, "STATE DESCRIPTIONS + VALUES  (SingleDevice §4.6)"
        )
        if st_desc:
            for name_key, entry in st_desc.items():
                row = _write_index_banner(ws, row, f"State  '{name_key}'")
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  description",
                    SCHEMA_STATE_DESC,
                    entry if isinstance(entry, dict) else {},
                )
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  current value",
                    SCHEMA_STATE_VAL,
                    st_vals.get(name_key, {}),
                )
        else:
            row = _write_schema_rows(ws, row, [("(empty)", "Required", "")], {})
        row += 1

    # ------------------------------------------------------------------ #
    # 13. Property Descriptions + Property Values
    # ------------------------------------------------------------------ #
    pr_desc = props.get("devicePropertyDescriptions")
    pr_vals = props.get("deviceProperties", {})
    if pr_desc is not None:
        row = _write_section_header(
            ws, row, "PROPERTY DESCRIPTIONS + VALUES  (SingleDevice §4.6)"
        )
        if pr_desc:
            for name_key, entry in pr_desc.items():
                current_val = pr_vals.get(name_key, _SENTINEL)
                row = _write_index_banner(
                    ws, row, f"Property  '{name_key}'  =  {_format_val(current_val)}"
                )
                row = _write_nested_schema_rows(
                    ws,
                    row,
                    "  description",
                    SCHEMA_PROP_DESC,
                    entry if isinstance(entry, dict) else {},
                )
        else:
            row = _write_schema_rows(ws, row, [("(empty)", "Required", "")], {})
        row += 1

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Build devices (sync — no network connection needed)
# ---------------------------------------------------------------------------


def build_sync():
    host = VdcHost(
        port=8444,
        model="pyDSvDCAPI Full Showcase",
        name="full-showcase-host",
        vendor_name=VENDOR_NAME,
        state_path=STATE_FILE,
    )
    vdc = Vdc(
        host=host,
        implementation_id="x-pydsvdcapi-full-showcase",
        name="Full Showcase VDC",
        model="pydsvdcapi-showcase",
        capabilities=VdcCapabilities(
            metering=False,
            identification=True,
            dynamic_definitions=True,
        ),
    )
    host.add_vdc(vdc)
    devices = build_all_devices(vdc)
    return vdc, devices


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    print("Building devices…")
    vdc, devices = build_sync()
    print(f"  → {len(devices)} devices built")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    vdc_props = vdc.get_properties()
    print("  VDC properties:", list(vdc_props))
    build_vdc_sheet(wb, vdc_props)

    for di in devices:
        v = di.vdsd
        props = v.get_properties()
        pg = int(v._primary_group) if v._primary_group is not None else 0
        print(f"  D{di.idx:02d}  {di.name}  ({len(props)} top-level keys)")
        build_vdsd_sheet(wb, di.idx, di.name, props, pg)

    wb.save(OUT_PATH)
    print(f"\nWritten: {OUT_PATH}")


if __name__ == "__main__":
    main()
